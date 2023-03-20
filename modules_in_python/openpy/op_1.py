# openpyxl - para trabalha com arquivos Excel, xlsx, xlsm, xltx e xltm
# Com eesa biblioteca será possível ler e escrever dados em célular
# específicas, formatar célular, inserir gráficos, criar fo´rmulas, adcionar imagens e outros elementos gráficos às sua planilhas
# Excel, como a criação de relatórios eanálise de dados e/ou facilitando a manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readhedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
worksheet: Worksheet = workbook.active  # type: ignore

# criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome        idade nota
    ['João',       14, 5.5],
    ['MAria',      13, 9.7],
    ['Luiz',       15, 8.8],
    ['Alberto',    16, 10],
]


# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)
workbook.save(WORKBOOK_PATH)
